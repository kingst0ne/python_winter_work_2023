'''
Есть Excel- файл, в первом листе которого находится выработка каждого программиста
В первой строке ФИО, во второй - результат работы за день

'''


import openpyxl
dct = {}
wb = openpyxl.load_workbook('test1.xlsx')

for listname in wb.sheetnames:
    ws = wb[listname]
    for i in range(ws.max_row):
        if dct.get(ws.cell(row=i+1,column=1).value):
            dct[ws.cell(row=i+1,column=1).value] += ws.cell(row=i+1,column=2).value
        else:
            dct[ws.cell(row=i + 1, column=1).value] = ws.cell(row=i + 1, column=2).value

if 'NewList' in wb.sheetnames:
    wb.remove('NewList')
wb.create_sheet('NewList')
ws = wb['NewList']

for i,key in enumerate(sorted(dct.items(),key=lambda x:x[0])):
    (ws.cell(row=i+1,column=1).value) = key[0]
    (ws.cell(row=i + 1, column=2).value) = key[1]


wb.save('test1.xlsx')