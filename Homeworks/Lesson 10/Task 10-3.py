'''

'''

import statistics as st
import openpyxl
dct = {}
wb = openpyxl.load_workbook('test2.xlsx')

for listname in wb.sheetnames:
    ws = wb[listname]
    for i in range(ws.max_row):
        if dct.get(ws.cell(row=i+1,column=1).value):
            dct[ws.cell(row=i+1,column=1).value] += ws.cell(row=i+1,column=2).value
        else:
            dct[ws.cell(row=i + 1, column=1).value] = ws.cell(row=i + 1, column=2).value

if 'NewList' in wb.sheetnames:
    wb.remove('NewList')
wb.create_sheet('NewList')
ws = wb['NewList']

ws.cell(row=1,column=1).value = 'Максимум'
ws.cell(row=1,column=2).value = max(dct.values())
ws.cell(row=2,column=1).value = 'Минимум'
ws.cell(row=2,column=2).value = min(dct.values())
ws.cell(row=3,column=1).value = 'Среднее арифметическое'
ws.cell(row=3,column=2).value = st.mean(dct.values())
ws.cell(row=4,column=1).value = 'Медиана'
ws.cell(row=4,column=2).value = st.median(dct.values())

wb.save('test2.xlsx')