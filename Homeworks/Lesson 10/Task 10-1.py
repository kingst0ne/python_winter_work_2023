'''
Есть Excel- файл, в первом листе которого находится выработка каждого программиста
В первой строке ФИО, во второй - результат работы за день
Дана информация за несколько дней по разным людям
На втором листе надо сформировать суммарные итоги по каждому сотруднику и общий итог

'''

import openpyxl
dct = {}
wb = openpyxl.load_workbook('test.xlsx')
ws = wb.active
if 'NewList' in wb.sheetnames:
    wb.remove('NewList')
wb.create_sheet('NewList')

for i in range(ws.max_row):
    if dct.get(ws.cell(row=i+1,column=1).value):
        dct[ws.cell(row=i+1,column=1).value] += ws.cell(row=i+1,column=2).value
    else:
        dct[ws.cell(row=i + 1, column=1).value] = ws.cell(row=i + 1, column=2).value
ws = wb['NewList']

for i,key in enumerate(dct):
    (ws.cell(row=i+1,column=1).value) = key
    (ws.cell(row=i + 1, column=2).value) = dct[key]

(ws.cell(row=len(dct)+1,column=1).value) = 'ИТОГО'
(ws.cell(row=len(dct) + 1, column=2).value) = sum(dct.values())

wb.save('test.xlsx')

pass