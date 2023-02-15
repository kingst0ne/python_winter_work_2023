'''
Дан файл с расширением .csv, содержащий в каждой строчке
следующую информацию - номер, фамилия, имя, компания, зп

Создать excel-файл, в который перенесите эту информацию,
предварительно отсортировав этот список по компании,
фамилии и имени

В конце списка добавьте строку ИТОГО и суммарное значение зарплат
'''

import csv
import openpyxl


lst = []
with open('data2.csv', newline='', encoding='utf-8') as file:
    rows = csv.reader(file)
    for i, row in enumerate (rows):
        if i == 0 and row[0].isalpha() or i == 0 and row[0].split(';')[0].isalpha():
            continue
        if len(row) > 1:
            lst.append(tuple(row))
        else:
            lst.append(tuple(row[0].split(';')))


lst_sorted = sorted(lst, key= lambda x: (x[3],x[1],x[2]))



wb = openpyxl.Workbook()
wb.save('salary.xlsx')
ws = wb.active
summa = 0
for i in range(1, len(lst_sorted)+1):
    for j in range(1, len(lst_sorted[0])+1):
        ws.cell(row=i, column=j).value = lst_sorted[i-1][j-1]
        if j == 5:
            summa += int(ws.cell(row=i, column=5).value)

ws.cell(row=i+1, column=1).value = 'ИТОГО'
ws.cell(row=i+1, column=5).value = summa

wb.save('salary.xlsx')



pass
